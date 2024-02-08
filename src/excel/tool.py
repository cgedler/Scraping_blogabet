# Created on : Feb 5, 2024, 7:11:04 PM
# Author     : Christopher Gedler


import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


def ReadExcelFile(fileName, fileRowNumbers):
    try:
        list =[]
        path = cwd = os.getcwd()
        fullPath = path + fileName
        workbook = openpyxl.load_workbook(fullPath, data_only=True)        
        sheet = workbook['Sheet1']
        for i in range(fileRowNumbers):
            y = i + 1
            valueRowColA = sheet['A' + '%d' % y].value
            value = str(valueRowColA)
            x = re.search("^https.*com", value)
            if x is not None:            
                result = x.group()
                list.append(result)
        return list                            
    except Exception as ex:
        print("Something went wrong when reading the file!!!! " + str(ex))

def writeExcelFile(fileName, listObj):
    try:
        wb = Workbook()
        ws = wb.active
        ws['B2'] = 'URL'
        ws['C2'] = 'Picks'
        ws['D2'] = 'Profit'
        ws['E2'] = 'Yields'
        ws['F2'] = 'Followers'
        for cell in ws[2]:
            cell.font = Font(size=10, bold=True)
        j = 3
        for i in listObj:
            ws['B' + '%d' % j] = str(i._url)
            ws['C' + '%d' % j] = str(i._picks)
            ws['D' + '%d' % j] = str(i._profit)
            ws['E' + '%d' % j] = str(i._yields)
            ws['F' + '%d' % j] = str(i._followers)
            j += 1 
        wb.save(fileName)
    except Exception as ex:
        print("Something went wrong when writing the file!!!! " + str(ex))